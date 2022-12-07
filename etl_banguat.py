import openpyxl
import pandas as pd
import locale
from datetime import datetime as dt
from dateutil.relativedelta import relativedelta
import os
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)

locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')


def isNumeric(valor):
    try:
        num = int(valor)
        return True
    except:
        try:
            num = float(valor)
            return True
        except:
            return False

#===================================
#**************BANGUAT**************
#===================================

#Definiendo dataframes

banguat_exportaciones_importaciones = pd.DataFrame({"reporte_nombre" : pd.Series(dtype = "str"), 
                                            "fecha" : pd.Series(dtype = "str"), 
                                            "tipo" : pd.Series(dtype = "str"), 
                                            "producto" : pd.Series(dtype = "str"), 
                                            "valor" : pd.Series(dtype = "float"), 
                                            "dimensional" : pd.Series(dtype = "str")})

banguat_divisa = pd.DataFrame({"reporte_nombre" : pd.Series(dtype = "str"), 
                                            "fecha" : pd.Series(dtype = "str"), 
                                            "concepto" : pd.Series(dtype = "str"), 
                                            "tipo_valor" : pd.Series(dtype = "str"), 
                                            "valor" : pd.Series(dtype = "float"), 
                                            "dimensional" : pd.Series(dtype = "str")})

banguat_monetario = pd.DataFrame({"reporte_nombre" : pd.Series(dtype = "str"), 
                                            "fecha" : pd.Series(dtype = "str"), 
                                            "valor" : pd.Series(dtype = "float"), 
                                            "dimensional" : pd.Series(dtype = "str")})



def monetario(archivo, offset = 0, remover_titulo = 0):
    global banguat_monetario

    print('Parsing', archivo + '.xlsx')

    imm = openpyxl.load_workbook("BANGUAT\\" + archivo + ".xlsx")
    años = []
    sheet = imm.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 2 + offset and col == 2:
                reporte_nombre = cell.value[:-remover_titulo]
            
            if row == 4 + offset and col == 2:
                dimensional = cell.value

            #Obteniendo años

            if row == 5 + offset and col >= 4 and not cell.value is None:
                años.append(str(cell.value)[0:4])
                
            #Obteniendo mes
            if row >= 6 + offset and col == 3 and not cell.value is None:
                mes = cell.value
                
            #Obteniendo valores
            if row >= 6 + offset and col >= 4 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value            
                fecha_valor = "1 de " + mes + " de " + años[col - 4]
                fecha = dt.strptime(fecha_valor, "%d de %B de %Y")
                banguat_monetario = banguat_monetario.append({"reporte_nombre" : reporte_nombre, 
                                            "fecha" : fecha, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)


def ExecuteETL():

    global banguat_exportaciones_importaciones
    global banguat_divisa
    global banguat_monetario

    print('=============================')
    print('****Executing BANGUAT ETL****')
    print('=============================')


    #dat1 - Exportaciones e importaciones
    
    print('Parsing data1.xlsx')

    impExp = openpyxl.load_workbook("BANGUAT\data1.xlsx")
    producto = ""
    tipo = ""
    año = "2022"
    meses = []
    tipos = []
    reporte_nombre = "VALOR (CIF) DE LAS IMPORTACIONES Y VALOR (FOB) DE LAS EXPORTACIONE POR PRODUCTO"
    dimensional = "En US Dólares"

    sheet = impExp.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            #Obteniendo tipo
            if row == 2 and col >= 4 and not cell.value is None:
                tipos.append(cell.value)

            #Obteniendo meses
            if row == 3 and col >= 4 and not cell.value is None:
                meses.append(cell.value)

            if row >= 5 and col == 2:
                producto = cell.value

            #Obteniendo valores
            if row >= 5 and col >= 4 and not cell.value is None and isNumeric(cell.value) and meses[col - 4] != "TOTAL":
                fecha_valor = "1 DE " + meses[col - 4] + " DE " + año
                fecha = dt.strptime(fecha_valor, "%d DE %B DE %Y")
                valor = cell.value
                
                banguat_exportaciones_importaciones = banguat_exportaciones_importaciones.append({"reporte_nombre" : reporte_nombre, 
                                            "fecha" : fecha, 
                                            "tipo" : tipos[col - 4],
                                            "producto" : producto, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)




    #gra051 - Oferta y demanda del mercado institucional divisas

    print('Parsing gra051.xlsx')

    g51 = openpyxl.load_workbook("BANGUAT\gra051.xlsx")
    concepto = ""
    tipo_valor = ""
    sheet = g51.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value
            
            if row == 2 and col == 1:
                dimensional = cell.value

            if row == 5 and col == 2:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-5:])
                fecha = dt.strptime(fecha_valor, "%d DE %B DE %Y")

            #Obteniendo concepto
            if row >= 9 and col == 1 and not cell.value is None:
                concepto = cell.value
            
            #Obteniendo valores
            if row >= 9 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                fecha_registro = fecha
                valor = cell.value
                if col == 2:
                    tipo_valor = "Semanal"
                if col == 3:
                    tipo_valor = "Acumulado"
                if col == 4: 
                    tipo_valor = "Acumulado"
                    fecha_registro = fecha - relativedelta(years = 1)
                if col == 5: 
                    tipo_valor = "Acumulado"
                    fecha_registro = fecha - relativedelta(years = 2)
                if col == 6:
                    tipo_valor = "Variación absoluta"
                if col == 7:
                    tipo_valor = "Variación relativa"
                if col == 8:
                    tipo_valor = "Variación absoluta"
                    fecha_registro = fecha - relativedelta(years = 1)
                if col == 9:
                    tipo_valor = "Variación relativa"
                    fecha_registro = fecha - relativedelta(years = 1)

                
                banguat_divisa = banguat_divisa.append({"reporte_nombre" : reporte_nombre, 
                                            "fecha" : fecha_registro, 
                                            "concepto" : concepto,
                                            "tipo_valor" : tipo_valor, 
                                            "valor" : float(valor), 
                                            "dimensional" : dimensional}, ignore_index = True)

    #gra052 - Ingreso de divisas por exportaciones

    print('Parsing gra052.xlsx')

    g52 = openpyxl.load_workbook("BANGUAT\gra052.xlsx")
    concepto = ""
    tipo_valor = ""
    sheet = g52.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value
            
            if row == 3 and col == 1:
                dimensional = cell.value

            if row == 6 and col == 2:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-5:])
                fecha = dt.strptime(fecha_valor, "%d DE %B DE %Y")

            #Obteniendo concepto
            if row >= 10 and col == 1 and not cell.value is None:
                concepto = cell.value
            
            #Obteniendo valores
            if row >= 10 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                fecha_registro = fecha
                valor = cell.value
                if col == 2:
                    tipo_valor = "Semanal"
                if col == 4:
                    tipo_valor = "Acumulado"
                if col == 5: 
                    tipo_valor = "Acumulado"
                    fecha_registro = fecha - relativedelta(years = 1)
                if col == 6:
                    tipo_valor = "Variación absoluta"
                if col == 7:
                    tipo_valor = "Variación relativa"
                
                banguat_divisa = banguat_divisa.append({"reporte_nombre" : reporte_nombre, 
                                            "fecha" : fecha_registro, 
                                            "concepto" : concepto,
                                            "tipo_valor" : tipo_valor, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)


    #gra053 - Ingreso de divisas por exportaciones

    print('Parsing gra053.xlsx')

    g53 = openpyxl.load_workbook("BANGUAT\gra053.xlsx")
    concepto = ""
    tipo_valor = ""
    sheet = g53.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value
            
            if row == 2 and col == 1:
                dimensional = cell.value

            if row == 5 and col == 2:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-5:])
                fecha = dt.strptime(fecha_valor, "%d DE %B DE %Y")

            #Obteniendo concepto
            if row >= 9 and col == 1 and not cell.value is None:
                concepto = cell.value
            
            #Obteniendo valores
            if row >= 9 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value
                if col == 2:
                    tipo_valor = "Ingresos Semanal"
                if col == 3:
                    tipo_valor = "Egresos Semanal"
                if col == 4:
                    tipo_valor = "Saldo Semanal"
                if col == 5: 
                    tipo_valor = "Ingresos Acumulado"
                if col == 6:
                    tipo_valor = "Egresos Acumulado"
                if col == 7:
                    tipo_valor = "Saldo Acumulado"
                
                banguat_divisa = banguat_divisa.append({"reporte_nombre" : reporte_nombre, 
                                            "fecha" : fecha, 
                                            "concepto" : concepto,
                                            "tipo_valor" : tipo_valor, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)









    monetario("imm07", 0, 3)
    monetario("imm08", 1, 3)
    monetario("imm12", -1, 6)
    monetario("imm13b", -1, 6)
    monetario("imm14", 1, 3)
    monetario("imm15", 0, 3)
    monetario("imm17", -1, 6)
    monetario("imm18a", -1, 6)
    monetario("imm20a", -1, 6)




    #GENERANDO ARCHIVOS PARA BASE DE DATOS

    print('Generating Excel files')

    def generateExcel(df, file, index):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        print('Generating', file)
        full_path = os.path.join(base_dir, file)
        if os.path.exists(full_path):
            print('File exists. Merging')
            df_old = pd.read_excel(file, index_col=False)
            df_new = pd.concat([df, df_old]).groupby(index, as_index = False).last()
            df_new.to_excel(file, index=False)
        else:
            df.to_excel(file, index=False)


    generateExcel(banguat_divisa, "BASES_DE_DATOS\BANGUAT_Divisa.xlsx", ['reporte_nombre', 'fecha', 'concepto', 'tipo_valor', 'dimensional'])
    generateExcel(banguat_monetario, "BASES_DE_DATOS\BANGUAT_Monetario.xlsx", ['reporte_nombre', 'fecha', 'dimensional'])
    generateExcel(banguat_exportaciones_importaciones, "BASES_DE_DATOS\\BANGUAT_Exportaciones_Importaciones.xlsx", ['reporte_nombre', 'fecha', 'tipo', 'producto', 'dimensional'])

    print('ETL for BANGUAT files completed!')

