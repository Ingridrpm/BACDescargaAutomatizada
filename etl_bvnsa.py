import openpyxl
import pandas as pd
import locale
from datetime import datetime as dt
from dateutil.relativedelta import relativedelta
import os
from xls2xlsx import XLS2XLSX
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)

locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')

def isNumeric(valor):
    try:
        num = int(valor)
        return True
    except:
        try:
            num = float(valor)
            return True
        except:
            return False


#===================================
#**************BVNSA**************
#===================================

#Convirtiendo archivos xls a xlsx

def convertToXslx(file): 
    print('Reconfiguring', file, 'xlsx file...')

    base_dir = os.path.dirname(os.path.abspath(__file__))
    full_path = os.path.join(base_dir, file)
    new_full_path = os.path.join(base_dir, file[:-1])

    if os.path.exists(new_full_path):
        os.remove(new_full_path)

    os.rename(full_path, full_path[:-1])

    x2x = XLS2XLSX(file[:-1])
    x2x.to_xlsx(file)
    
    print('File converted:', full_path)


def generateExcel(df, file):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        print('Generating', file)
        full_path = os.path.join(base_dir, file)
        if os.path.exists(full_path):
            print('File exists. Merging')
            df_old = pd.read_excel(file, index_col=False)
            df_new = pd.concat([df, df_old]).drop_duplicates()
            df_new.to_excel(file, index=False)
        else:
            df.to_excel(file, index=False)


def bolsa_valores_tabular(file):

    #convertToXslx(file)

    print('Parsing', file)

    bv = openpyxl.load_workbook(file)
    sheet = bv.active
    
    df = pd.DataFrame()

    titulo = str(sheet.cell(row = 2, column = 2).value)
    reporte_nombre = " ".join(titulo.split()[:-2]).replace(' ', '_')
    fecha_valor = " ".join(titulo.split()[-1:])
    fecha = dt.strptime(fecha_valor, "%d/%m/%Y")

    for col in range(1, sheet.max_column + 1):
        valores = []
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row = row, column = col)

            #Obteniendo características
            if row == 5 and not cell.value is None:
                caracteristica = str(cell.value).lower().replace(' ', '_')
                
            #Obteniendo valores
            if row >= 6:
                valores.append(cell.value)

        df[caracteristica] = valores

    df.insert(0, 'fecha_reporte', fecha)
    
    generateExcel(df, ".\\BASES_DE_DATOS\\BVNSA_" + reporte_nombre + ".xlsx")


def bolsa_valores_tabular2(file):

    convertToXslx(file)

    print('Parsing', file + '.xlsx')

    bv = openpyxl.load_workbook(".\\BVNSA\\" + file + ".xlsx")
    sheet = bv.active
    
    df = pd.DataFrame()

    titulo = str(sheet.cell(row = 2, column = 2).value)
    reporte_nombre = file.replace(' ', '_')
    fecha_valor = bv.sheetnames[0]
    fecha = dt.strptime(fecha_valor, "%d-%m-%Y")
    
    for col in range(1, sheet.max_column + 1):
        valores = []
        caracteristica = ''

        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row = row, column = col)

            #Obteniendo características
            if row == 1 and col >= 2:
                caracteristica = str(cell.value).lower().replace(' ', '_')
                
            #Obteniendo valores
            if row >= 2 and col >= 2:
                valores.append(cell.value)

        if caracteristica != '':
            df[caracteristica] = valores

    df.insert(0, 'fecha_reporte', fecha)
    
    generateExcel(df, ".\\BASES_DE_DATOS\\BVNSA_" + reporte_nombre + ".xlsx")


def ExecuteETL():

    print('=============================')
    print('*****Executing BVNSA ETL*****')
    print('=============================')

    bolsa_valores_tabular('.\\BVNSA\\CurvaNss.xlsx')
    bolsa_valores_tabular('.\\BVNSA\\CurvaRendimiento.xlsx')
    bolsa_valores_tabular('.\\BVNSA\\Informe_Reportos_Privados.xlsx')
    bolsa_valores_tabular('.\\BVNSA\\Informe_Reportos_Publicos.xlsx')
    bolsa_valores_tabular('.\\BVNSA\\Informe_Resumen_Reportos.xlsx')
    bolsa_valores_tabular('.\\BVNSA\\OperacionesSinedi.xlsx')
    bolsa_valores_tabular2('.\\BVNSA\\Informe Diario.xlsx')

    print('ETL for BVNSA files completed!')

