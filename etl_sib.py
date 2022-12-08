import openpyxl
import pandas as pd
import locale
from datetime import datetime as dt
import os
import warnings
from xls2xlsx import XLS2XLSX


warnings.simplefilter(action='ignore', category=FutureWarning)

locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')

#Definiendo dataframe

sib_informacion_bancaria = pd.DataFrame({"reporte_nombre" : pd.Series(dtype = "str"), 
                                            "entidad" : pd.Series(dtype = "str"), 
                                            "fecha" : pd.Series(dtype = "str"), 
                                            "tipo_moneda" : pd.Series(dtype = "str"), 
                                            "institucion_nombre" : pd.Series(dtype = "str"), 
                                            "caracteristica" : pd.Series(dtype = "str"), 
                                            "valor" : pd.Series(dtype = "float"), 
                                            "dimensional" : pd.Series(dtype = "str")})

sib_actividad_economica = pd.DataFrame({"reporte_nombre" : pd.Series(dtype = "str"), 
                                            "fecha" : pd.Series(dtype = "str"), 
                                            "tipo_moneda" : pd.Series(dtype = "str"), 
                                            "actividad_economica" : pd.Series(dtype = "str"), 
                                            "caracteristica" : pd.Series(dtype = "str"), 
                                            "valor" : pd.Series(dtype = "float"), 
                                            "dimensional" : pd.Series(dtype = "str")})




def isNumeric(valor):
    try:
        num = int(valor)
        return True
    except:
        try:
            num = float(valor)
            return True
        except:
            return False

#===================================
#***************SIB*****************
#===================================

#Convirtiendo archivos xls a xlsx

def convertToXslx(file): 
    print('Converting', file, 'to xlsx...')

    base_dir = os.path.dirname(os.path.abspath(__file__))
    full_path = os.path.join(base_dir, file)
    new_full_path = os.path.join(base_dir, file + 'x')

    if os.path.exists(new_full_path):
        os.remove(new_full_path)

    x2x = XLS2XLSX(file)
    x2x.to_xlsx(file + 'x')
    
    
    print('File converted:', full_path)

def ExecuteETL():

    global sib_informacion_bancaria
    global sib_actividad_economica

    print('=============================')
    print('******Executing SIB ETL******')
    print('=============================')

    convertToXslx('..\\SIB\\Consulta_7.xls')
    convertToXslx('..\\SIB\\Consulta_18.xls')
    convertToXslx('..\\SIB\\Consulta_19.xls')
    convertToXslx('..\\SIB\\Consulta_20.xls')
    convertToXslx('..\\SIB\\Consulta_21.xls')
    convertToXslx('..\\SIB\\Consulta_26.xls')
    convertToXslx('..\\SIB\\Consulta_198.xls')
    convertToXslx('..\\SIB\\Consulta_200.xls')
    convertToXslx('..\\SIB\\Consulta_205.xls')
    convertToXslx('..\\SIB\\Consulta_210.xls')
    convertToXslx('..\\SIB\\Consulta_383.xls')
    convertToXslx('..\\SIB\\Consulta_384.xls')
    convertToXslx('..\\SIB\\Consulta_412.xls')


    reporte_nombre = ""
    entidad = ""
    fecha = ""
    tipo_moneda = ""
    institucion_nombre = ""
    caracteristica = ""
    valor = 0.0
    dimensional = ""

    actividad_economica = ""

    #Consulta_7 - Balance y estado de resultados

    print('Parsing Consulta_7.xlsx')

    c7 = openpyxl.load_workbook("..\\SIB\\Consulta_7.xlsx")
    entidad = "Local"
    tipo_moneda = "Moneda nacional"
    instituciones_bancarias = []

    sheet = c7.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row == 3 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-5:])
                fecha = dt.strptime(fecha_valor, "%d de %B de %Y")

            if row == 4 and col == 1:
                dimensional = cell.value

            #Obteniendo instituciones bancarias
            
            if row == 5 and col >= 2:
                instituciones_bancarias.append(cell.value)

            #Obteniendo características
            if row >= 6 and col == 1 and not cell.value is None:
                caracteristica = cell.value
            
            #Obteniendo valores
            if row >= 6 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : instituciones_bancarias[col - 2], 
                                            "caracteristica" : caracteristica, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)
            
    #Consulta_18 - Tasas de interés aplicadas

    print('Parsing Consulta_18.xlsx')

    c18 = openpyxl.load_workbook("..\\SIB\\\\Consulta_18.xlsx")
    entidad = "Local"
    caracteristicas = []
    dimensional = "Porcentaje"
    seccion = 1
    moneda = "Moneda nacional"

    sheet = c18.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row > 1 and seccion == 1 and cell.value == reporte_nombre:
                seccion += 1
                tipo_moneda = "Moneda extranjera"

            if row == 3 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-1:])
                fecha = dt.strptime(fecha_valor, "%d/%m/%Y")

            #Obteniendo caracteristicas
            
            if row == 4 and col >= 2:
                caracteristicas.append(str(cell.value).replace('\r', '').replace('\n', ' '))

            #Obteniendo instituciones bancarias
            if row >= 5 and col == 1 and not cell.value is None:
                if not cell.value.startswith("La tasa") and not cell.value.startswith("Las tasas") and cell.value != reporte_nombre and not cell.value.startswith("MONEDA") and not cell.value.startswith("INSTITUCIONES BANCARIAS") and not cell.value.startswith("Al "):
                    institucion_bancaria = cell.value
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : institucion_bancaria,
                                            "caracteristica" :  caracteristicas[col - 2], 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)


    #Consulta_19 - Tasas de interés aplicadas semanal

    print('Parsing Consulta_19.xlsx')

    c19 = openpyxl.load_workbook("..\\SIB\\\\Consulta_19.xlsx")
    entidad = "Local"
    caracteristicas = []
    dimensional = "Porcentaje"
    seccion = 1
    moneda = "Moneda nacional"

    sheet = c19.active

    for row in range(1, sheet.max_row + 1):
                
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row > 1 and seccion == 1 and cell.value == reporte_nombre:
                seccion += 1
                tipo_moneda = "Moneda extranjera"

            if row == 3 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-1:])
                fecha = dt.strptime(fecha_valor, "%d/%m/%Y")

            #Obteniendo caracteristicas
            
            if row == 4 and col >= 2:
                caracteristicas.append(str(cell.value).replace('\r', '').replace('\n', ' '))

            #Obteniendo instituciones bancarias
            if row >= 5 and col == 1 and not cell.value is None:
                if not cell.value.startswith("La tasa") and not cell.value.startswith("Las tasas") and cell.value != reporte_nombre and not cell.value.startswith("MONEDA") and not cell.value.startswith("INSTITUCIONES BANCARIAS") and not cell.value.startswith("Al "):
                    institucion_bancaria = cell.value
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : institucion_bancaria,
                                            "caracteristica" :  caracteristicas[col - 2], 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)


    #Consulta_20 - Tasas promedio ponderadas

    print('Parsing Consulta_20.xlsx')

    c20 = openpyxl.load_workbook("..\\SIB\\\\Consulta_20.xlsx")
    entidad = "Local"
    caracteristicas = []
    dimensional = "Porcentaje"
    seccion = 1
    moneda = "Moneda nacional"

    sheet = c20.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row > 1 and seccion == 1 and cell.value == reporte_nombre:
                seccion += 1
                tipo_moneda = "Moneda extranjera"

            if row == 3 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-1:])
                fecha = dt.strptime(fecha_valor, "%d/%m/%Y")

            #Obteniendo caracteristicas
            
            if row == 4 and col >= 2:
                caracteristicas.append(str(cell.value).replace('\r', '').replace('\n', ' '))

            #Obteniendo instituciones bancarias
            if row >= 5 and col == 1 and not cell.value is None:
                if not cell.value.startswith("MONEDA") and not cell.value.startswith("INSTITUCIONES BANCARIAS") and not cell.value.startswith("Al ") and cell.value != reporte_nombre:
                    institucion_bancaria = cell.value
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : institucion_bancaria,
                                            "caracteristica" :  caracteristicas[col - 2], 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)

    #Consulta_26 - Cartera de créditos por agrupación

    print('Parsing Consulta_26.xlsx')

    c26 = openpyxl.load_workbook("..\\SIB\\\\Consulta_26.xlsx")
    entidad = "Local"
    caracteristicas = []
    dimensional = "Quetzales"
    seccion = 1
    moneda = "Moneda nacional"

    sheet = c26.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row > 1 and seccion == 1 and cell.value == reporte_nombre:
                seccion += 1
                tipo_moneda = "Moneda extranjera"

            if row > 1 and seccion == 2 and cell.value == reporte_nombre:
                seccion += 1
                tipo_moneda = "Moneda local y extranjera"

            if row == 3 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-1:])
                fecha = dt.strptime(fecha_valor, "%d/%m/%Y")

            #Obteniendo caracteristicas
            
            if row == 4 and col >= 2:
                caracteristicas.append(cell.value)

            #Obteniendo instituciones bancarias
            if row >= 5 and col == 1 and not cell.value is None:
                if not cell.value.startswith("MONEDA") and not cell.value.startswith("INSTITUCIONES") and not cell.value.startswith("Al ") and cell.value != reporte_nombre:
                    institucion_bancaria = cell.value
                    if institucion_bancaria == "SISTEMA":
                        institucion_bancaria = "SISTEMA BANCARIO"
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : institucion_bancaria,
                                            "caracteristica" :  caracteristicas[col - 2], 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)



    #Consulta_198 - Balance y estado de resultados Off-Shore

    print('Parsing Consulta_198.xlsx')

    c198 = openpyxl.load_workbook("..\\SIB\\\\Consulta_198.xlsx")
    entidad = "Off-Shore"
    tipo_moneda = "Moneda nacional"
    instituciones_bancarias = []

    sheet = c198.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row == 2 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-5:])
                fecha = dt.strptime(fecha_valor, "%d de %B de %Y")

            if row == 3 and col == 1:
                dimensional = cell.value

            #Obteniendo instituciones bancarias
            
            if row == 4 and col >= 2:
                instituciones_bancarias.append(cell.value)

            #Obteniendo características
            if row >= 5 and col == 1 and not cell.value is None:
                caracteristica = cell.value
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : instituciones_bancarias[col - 2], 
                                            "caracteristica" : caracteristica, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)
            

    #Consulta_200 - Tasas de interés aplicadas off-shore

    print('Parsing Consulta_200.xlsx')

    c200 = openpyxl.load_workbook("..\\SIB\\\\Consulta_200.xlsx")
    entidad = "Off-shore"
    caracteristicas = []
    dimensional = "Porcentaje"
    tipo_moneda = "Moneda extranjera"
    fecha = ""

    sheet = c200.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value.upper()

            #Obteniendo caracteristicas
            
            if row == 4 and col >= 2 and not cell.value is None:
                caracteristicas.append(str(cell.value).replace('\r', '').replace('\n', ' '))

            #Obteniendo instituciones bancarias
            if row >= 5 and col == 1 and not cell.value is None:
                institucion_bancaria = cell.value
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : institucion_bancaria,
                                            "caracteristica" :  caracteristicas[col - 2], 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)


    #Consulta_205 - Cartera de créditos por agrupación off-shore

    print('Parsing Consulta_205.xlsx')

    c205 = openpyxl.load_workbook("..\\SIB\\\\Consulta_205.xlsx")
    entidad = "Off-shore"
    caracteristicas = []
    dimensional = "Quetzales"
    seccion = 1
    
    sheet = c205.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row > 1 and seccion == 1 and cell.value == reporte_nombre:
                seccion += 1
                tipo_moneda = "Moneda extranjera"

            if row > 1 and seccion == 2 and cell.value == reporte_nombre:
                seccion += 1
                tipo_moneda = "Moneda local y extranjera"

            if row == 3 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-1:])
                fecha = dt.strptime(fecha_valor, "%d/%m/%Y")

            #Obteniendo caracteristicas
            
            if row == 4 and col >= 2:
                caracteristicas.append(cell.value)

            #Obteniendo instituciones bancarias
            if row >= 5 and col == 1 and not cell.value is None:
                if not cell.value.startswith("MONEDA") and not cell.value.startswith("INSTITUCIONES") and not cell.value.startswith("Al ") and cell.value != reporte_nombre:
                    institucion_bancaria = cell.value
                    if institucion_bancaria == "SISTEMA":
                        institucion_bancaria = "SISTEMA BANCARIO"
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : institucion_bancaria,
                                            "caracteristica" :  caracteristicas[col - 2], 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)


    #Consulta_210 - Balance y estado de resultados empresas especializadas

    print('Parsing Consulta_210.xlsx')

    c210 = openpyxl.load_workbook("..\\SIB\\\\Consulta_210.xlsx")
    entidad = "Empresa especializada"
    tipo_moneda = "Moneda nacional"
    instituciones_bancarias = []

    sheet = c210.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row == 2 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-5:])
                fecha = dt.strptime(fecha_valor, "%d de %B de %Y")

            if row == 3 and col == 1:
                dimensional = cell.value

            #Obteniendo instituciones bancarias
            
            if row == 4 and col >= 2:
                instituciones_bancarias.append(cell.value)

            #Obteniendo características
            if row >= 5 and col == 1 and not cell.value is None:
                caracteristica = cell.value
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : instituciones_bancarias[col - 2], 
                                            "caracteristica" : caracteristica, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)


    #Consulta_383 Perfil financiero

    print('Parsing Consulta_383.xlsx')

    c383 = openpyxl.load_workbook("..\\SIB\\\\Consulta_383.xlsx")
    entidad = "Local"
    tipo_moneda = "Moneda nacional"
    instituciones_bancarias = []

    sheet = c383.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 2 and col == 1:
                reporte_nombre = cell.value

            if row == 3 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-5:])
                fecha = dt.strptime(fecha_valor, "%d de %B de %Y")

            if row == 4 and col == 1:
                dimensional = cell.value

            #Obteniendo instituciones bancarias
            
            if row == 5 and col >= 2:
                instituciones_bancarias.append(cell.value)

            #Obteniendo características
            if row >= 6 and col == 1 and not cell.value is None:
                caracteristica = cell.value
            
            #Obteniendo valores
            if row >= 6 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : instituciones_bancarias[col - 2], 
                                            "caracteristica" : caracteristica, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)


    #Consulta_412 - Cartera de créditos de generadores y no generadores de divisas

    print('Parsing Consulta_412.xlsx')

    c412 = openpyxl.load_workbook("..\\SIB\\\\Consulta_412.xlsx")
    entidad = "Local"
    caracteristicas = []
    tipo_moneda = "Moneda nacional"

    sheet = c412.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if row == 2 and col == 1:
                reporte_nombre = cell.value

            if row == 3 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-1:])
                fecha = dt.strptime(fecha_valor, "%d/%m/%Y")

            if row == 4 and col == 1:
                dimensional = cell.value

            #Obteniendo caracteristicas
            
            if row == 5 and col >= 2:
                caracteristicas.append(str(cell.value).replace('\r', '').replace('\n', ' '))

            #Obteniendo instituciones bancarias
            if row >= 6 and col == 1 and not cell.value is None:
                institucion_bancaria = cell.value
            
            #Obteniendo valores
            if row >= 6 and col >= 2 and not cell.value is None and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : institucion_bancaria,
                                            "caracteristica" :  caracteristicas[col - 2], 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)

    #Consulta_384 - Indicadores financieros

    c384 = openpyxl.load_workbook("..\\SIB\\\\Consulta_384.xlsx")
    entidad = "Local"
    tipo_moneda = "Moneda nacional"
    dimensional = "Porcentaje"
    stop = False

    sheet = c384.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if cell.value == "AREA":
                stop = True

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row == 2 and col == 1:
                institucion_bancaria = cell.value

            if row == 4 and col == 2:
                fecha_valor = cell.value
                
            #Obteniendo características
            if row >= 5 and col == 1 and not cell.value is None and not stop:
                caracteristica = cell.value
            
            #Obteniendo valores
            if row >= 5 and col >= 2 and not cell.value is None and not stop and isNumeric(cell.value):
                valor = cell.value 
                sib_informacion_bancaria = sib_informacion_bancaria.append({"reporte_nombre" : reporte_nombre, 
                                            "entidad": entidad,
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "institucion_nombre" : institucion_bancaria,
                                            "caracteristica" : caracteristica, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)




    ##################Actividad económica###########################

    #Consulta_21 - Tasas de interés aplicadas

    c21 = openpyxl.load_workbook("..\\SIB\\\\Consulta_21.xlsx")
    entidad = "Local"
    actividades_economicas = []
    dimensional = "Porcentaje"
    seccion = 1
    moneda = "Moneda nacional"
    stop = False

    sheet = c21.active

    for row in range(1, sheet.max_row + 1):
        
        for col in range(1, sheet.max_column + 1):
            
            cell = sheet.cell(row = row, column = col)

            if cell.value == "Notas:":
                stop = True

            if row == 1 and col == 1:
                reporte_nombre = cell.value

            if row > 1 and seccion == 1 and cell.value == reporte_nombre:
                seccion += 1
                tipo_moneda = "Moneda extranjera"

            if row == 4 and col == 1:
                fecha_valor = cell.value
                fecha_valor = " ".join(fecha_valor.split()[-1:])
                fecha = dt.strptime(fecha_valor, "%d/%m/%Y")

            #Obteniendo actividades económicas
            
            if row == 5 and col >= 2:
                codigo_actividad_economica = cell.value
                if codigo_actividad_economica == "'(0)":
                    actividad_economica = "(0) Consumo, Transferencias y Otros Destinos"
                elif codigo_actividad_economica == "'(1)":
                    actividad_economica = "(1) Agricultura, Ganadería, Silvicultura, Caza y Pesca"
                elif codigo_actividad_economica == "'(2)":
                    actividad_economica = "(2) Explotación de Minas y Cantera"
                elif codigo_actividad_economica == "'(3)":
                    actividad_economica = "(3) Industria Manufacturera"
                elif codigo_actividad_economica == "'(4)":
                    actividad_economica = "(4) Electricidad, Gas y Agua"
                elif codigo_actividad_economica == "'(5)":
                    actividad_economica = "(5) Construcción"
                elif codigo_actividad_economica == "'(6)":
                    actividad_economica = "(6) Comercio"
                elif codigo_actividad_economica == "'(7)":
                    actividad_economica = "(7) Transporte y Almacenamiento"
                elif codigo_actividad_economica == "'(8)":
                    actividad_economica = "(8) Establecimientos Financieros, Bienes Inmuebles y Servicios Prestados a las Empresas"
                elif codigo_actividad_economica == "'(9)":
                    actividad_economica = "(9) Servicios Comunales, Sociales y Personales"

                actividades_economicas.append(actividad_economica)

            #Obteniendo caracteristicas
            if row >= 6 and col == 1 and not cell.value is None and not stop:
                if not cell.value.startswith("POR ACTIVIDAD") and not cell.value.startswith("Las tasas") and cell.value != reporte_nombre and not cell.value.startswith("MONEDA") and not cell.value.startswith("MONTO DEL") and not cell.value.startswith("Al "):
                    caracteristica = cell.value
            
            #Obteniendo valores
            if row >= 6 and col >= 2 and not cell.value is None and not stop and isNumeric(cell.value):
                valor = cell.value 
                sib_actividad_economica = sib_actividad_economica.append({"reporte_nombre" : reporte_nombre, 
                                            "fecha" : fecha, 
                                            "tipo_moneda" : tipo_moneda, 
                                            "actividad_economica" : actividades_economicas[col - 2], 
                                            "caracteristica" : caracteristica, 
                                            "valor" : valor, 
                                            "dimensional" : dimensional}, ignore_index = True)

    #GENERANDO ARCHIVOS PARA BASE DE DATOS

    print('Generating Excel files...')

    def generateExcel(df, file, index):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        print('Generating', file)
        full_path = os.path.join(base_dir, file)
        if os.path.exists(full_path):
            print('File exists. Merging')
            df_old = pd.read_excel(file, index_col=False)
            df_new = pd.concat([df, df_old]).groupby(index, as_index = False).last()
            df_new.to_excel(file, index=False)
        else:
            df.to_excel(file, index=False)


    generateExcel(sib_actividad_economica, "..\\BASES_DE_DATOS\\SIB_Actividad_Económica.xlsx", ["reporte_nombre", "fecha", "tipo_moneda", "actividad_economica", "caracteristica", "dimensional"])
    generateExcel(sib_informacion_bancaria, "..\\BASES_DE_DATOS\\SIB_Información_Bancaria.xlsx", ["reporte_nombre", "entidad", "fecha", "tipo_moneda", "institucion_nombre", "caracteristica", "dimensional"])

    print('ETL for SIB files completed!')